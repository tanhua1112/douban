# -*- coding: utf-8 -*-

import sys
import requests
import urllib.request
import chardet      # 编码转化处理
import traceback    # 异常处理
import random
import numpy as np
from bs4 import BeautifulSoup
from lxml import etree
import time
from openpyxl import Workbook

# ip池、免费IP池不好用

BaseUrl = 'https://www.xicidaili.com/nn/1'

class IpPool(object):
    def __init__(self):
        self.headers = [{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},
                        {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
                        {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
                        {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36'}]
        print(random.choice(self.headers))

    def IP_random(self, url):
        print('获取代理IP: {}'.format(url))
        try:
            resp = requests.get(url, headers=random.choice(self.headers), timeout=1)
            if resp.status_code == 200:
                print('解析提取页面href数据: {}'.format(url))
                soup = BeautifulSoup(resp.text, 'lxml')
                ips = soup.find_all('tr')
                proxies = []
                for i in range(1, len(ips)):
                    ip_info = ips[i]
                    tds = ip_info.find_all('td')
                    proto = tds[5].text   # python3 不加encode
                    ip = tds[1].text  # python3 不加encode
                    port = tds[2].text    # python3 不加encode
                    proxy = {proto: proto+'://'+ip+':'+port}
                    # proxy = {proto: '{0:s}://{1:s}:{2:s}'.format(proto, ip, port)}
                    proxies.append(proxy)
                return self.headers, proxies
                    # ip_list.append(tds[1].text + ':' + tds[2].text)                    
            else:
                print('页面href数据提取失败: {}'.format(url))
                return [], []
        except Exception:
            print('页面数据获取失败: {}'.format(url))
            traceback.print_exc()
    

def label_get(url, header, proxie):
    print('获取界面数据: {}'.format(url))
    try:
        resp = requests.get(url, headers=header, proxies=proxie, timeout=1)
        resp.encoding = chardet.detect(resp.content)['encoding']
        page_list = []
        if resp.status_code == 200:
            print('获取页面数据')
            # 解析界面标签 并返回标签信息
            html = etree.HTML(resp.text)
            labels = html.xpath('//div[@class="indent tag_cloud"]/table/tbody/tr/td/a/text()')
            labels = list(map(lambda x:x.strip().encode('utf-8'), labels))
            return labels
        else:
            return []
    except Exception:
        print('页面数据请求失败: {}'.format(url))
        traceback.print_exc()


def book_spider(book_tag, header, proxie):
    page_num=0
    book_list=[]
    try_times=0
    # while(1):
    while(1):
        url='http://www.douban.com/tag/'+urllib.parse.quote(book_tag)+'/book?start='+str(page_num*20)
        time.sleep(np.random.rand()*5)
        randheader=random.choice(header)
        randproxie=random.choice(proxie)
        try:
            resp = requests.get(url, headers=randheader, proxies=randproxie, timeout=1)
            resp.encoding = chardet.detect(resp.content)['encoding']
            if resp.status_code == 200:
                print('解析页面数据{}'.format(url))
                html = etree.HTML(resp.text)
                titles = html.xpath('//div[@class="mod book-list"]/dl/dd/a/text()')
                titles = list(map(lambda d: d.strip(), titles)) # py3 不加encode
                links = html.xpath('//div[@class="mod book-list"]/dl/dd/a/@href')
                links = list(map(lambda d: d.strip(), links))   # py3 不加encode
                guards = html.xpath('//div[@class="mod book-list"]/dl/dd/div/span/text()')
                guards = list(map(lambda d: d.strip(), guards)) # py3 不加encode
                infos = html.xpath('//div[@class="mod book-list"]/dl/dd/div/text()')
                infos = list(map(lambda d: d.strip(), infos))   # py3 不加encode    infos = list(map(lambda d: d.strip().encode('utf-8'), infos))
                infos = list(filter(lambda d: bool(d), infos))
                if titles:
                    for book_num in range(len(infos)):
                        desc = infos[book_num].split('/')
                        try:
                            author_info = '作者/译者： ' + '/'.join(desc[0:-3])
                        except:
                            author_info ='作者/译者： 暂无'
                        try:
                            pub_info = '出版信息： ' + '/'.join(desc[-3:])
                        except:
                            pub_info = '出版信息： 暂无'
                        book_list.append([titles[book_num], links[book_num], guards[book_num], author_info, pub_info])
                else:
                    break
        except Exception as ex:
            print(ex)
            continue
        finally:
            page_num+=1
    return book_list


def do_spider(book_tag_lists, headers, proxies):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag, headers, proxies)
        book_list=sorted(book_list,key=lambda x:x[2],reverse=True)
        book_lists.append(book_list)
    return book_lists


def data_2_excel(book_lists,book_tag_lists):
    wb=Workbook(write_only=True)
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i]))    # utf8->unicode
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号','书名','链接','评分','作者','出版社'])
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],bl[1],float(bl[2]),bl[3],bl[4]])
            count+=1
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i])
    save_path+='.xlsx'
    wb.save(save_path)


if __name__ == '__main__':
    headers, proxies = IpPool().IP_random(BaseUrl)
    print(headers, proxies)
    book_tag_lists = label_get('https://book.douban.com/tag/?view=cloud', random.choice(headers), random.choice(proxies))[:1]
    book_tag_lists = ['商业']  
    book_lists = do_spider(book_tag_lists, headers, proxies)
    data_2_excel(book_lists, book_tag_lists)
    # book_tag_lists = ['商业','理财','管理', '商 业','理 财','管 理']  
